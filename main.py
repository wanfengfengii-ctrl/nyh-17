from fastapi import FastAPI, Depends, HTTPException, status, Request, Form, File, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_
from datetime import date, timedelta, datetime
from typing import Optional, List
import os
import uuid
from io import BytesIO

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from database import engine, Base, get_db
from models import (
    User, Pottery, CleaningRecord, PotteryGroup, PotteryGroupMember, StorageRecord,
    PotteryImage, RepairTask, GroupVersion, StorageApproval, OperationLog,
    TaskStatus, ApprovalStatus, OperationType, UnderwaterImage,
    RepairPlan, ReviewRecord, RepairComparisonImage, RepairPlanStatus, RepairProgress
)
from schemas import (
    PotteryCreate, PotteryUpdate, CleaningRecordCreate, PotteryGroupCreate, PotteryGroupUpdate,
    StorageRecordCreate, RepairTaskCreate, RepairTaskUpdate, StorageRecordUpdate,
    UnderwaterImageCreate, UnderwaterImageUpdate,
    RepairPlanCreate, RepairPlanUpdate, ReviewRecordCreate, ReviewRecordUpdate
)
from auth import get_password_hash, authenticate_user, create_access_token, ACCESS_TOKEN_EXPIRE_MINUTES, get_current_user

Base.metadata.create_all(bind=engine)

app = FastAPI(title="陶片修复协同与审批管理系统")

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

UPLOAD_DIR = "static/uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)


def datetime_format(value, format="%Y-%m-%d"):
    if value:
        return value.strftime(format)
    return ""


def datetime_format_full(value, format="%Y-%m-%d %H:%M"):
    if value:
        return value.strftime(format)
    return ""


templates.env.filters["strftime"] = datetime_format
templates.env.filters["strftime_full"] = datetime_format_full


def log_operation(db: Session, user: User, operation_type: str, target_type: str, target_id: int, description: str, ip_address: str = ""):
    log = OperationLog(
        user_id=user.id,
        operation_type=operation_type,
        target_type=target_type,
        target_id=target_id,
        description=description,
        ip_address=ip_address
    )
    db.add(log)
    db.commit()


def init_default_user(db: Session):
    admin = db.query(User).filter(User.username == "admin").first()
    if not admin:
        hashed_password = get_password_hash("admin123")
        admin = User(username="admin", hashed_password=hashed_password, full_name="系统管理员", role="管理员")
        db.add(admin)
        db.commit()
    
    restorer = db.query(User).filter(User.username == "restorer").first()
    if not restorer:
        hashed_password = get_password_hash("restorer123")
        restorer = User(username="restorer", hashed_password=hashed_password, full_name="张修复", role="修复员")
        db.add(restorer)
        db.commit()
    
    approver = db.query(User).filter(User.username == "approver").first()
    if not approver:
        hashed_password = get_password_hash("approver123")
        approver = User(username="approver", hashed_password=hashed_password, full_name="李审批", role="审批员")
        db.add(approver)
        db.commit()


with next(get_db()) as db:
    init_default_user(db)


@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})


@app.post("/login")
async def login(request: Request, username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    user = authenticate_user(db, username, password)
    if not user:
        return templates.TemplateResponse("login.html", {"request": request, "error": "用户名或密码错误"})
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(data={"sub": user.username}, expires_delta=access_token_expires)
    response = RedirectResponse(url="/", status_code=status.HTTP_303_SEE_OTHER)
    response.set_cookie(key="access_token", value=f"Bearer {access_token}", httponly=True)
    return response


@app.get("/logout")
async def logout():
    response = RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    response.delete_cookie("access_token")
    return response


@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    total_potteries = db.query(Pottery).count()
    total_cleaned = db.query(Pottery).filter(Pottery.current_status == "已清洗").count()
    total_groups = db.query(PotteryGroup).count()
    total_stored = db.query(StorageRecord).filter(StorageRecord.is_official == True).count()
    total_tasks = db.query(RepairTask).count()
    pending_tasks = db.query(RepairTask).filter(RepairTask.status == TaskStatus.PENDING).count()
    pending_approvals = db.query(StorageRecord).filter(StorageRecord.approval_status == ApprovalStatus.SUBMITTED).count()
    
    total_repair_plans = db.query(RepairPlan).count()
    pending_reviews = db.query(RepairPlan).filter(RepairPlan.status == RepairPlanStatus.IN_REVIEW).count()
    in_repair_plans = db.query(RepairPlan).filter(RepairPlan.status == RepairPlanStatus.IN_REPAIR).count()
    completed_repair_plans = db.query(RepairPlan).filter(RepairPlan.status == RepairPlanStatus.COMPLETED).count()
    
    total_underwater_images = db.query(UnderwaterImage).count()
    total_3d_models = db.query(UnderwaterImage).filter(UnderwaterImage.file_type == "3d").count()
    total_photos = db.query(UnderwaterImage).filter(UnderwaterImage.file_type == "image").count()
    
    underwater_image_stats = {
        "total": total_underwater_images,
        "photos": total_photos,
        "models": total_3d_models,
        "by_water_area": {}
    }
    water_areas = [r[0] for r in db.query(UnderwaterImage.water_area).distinct().all() if r[0]]
    for area in water_areas:
        underwater_image_stats["by_water_area"][area] = db.query(UnderwaterImage).filter(
            UnderwaterImage.water_area == area
        ).count()

    recent_potteries = db.query(Pottery).order_by(Pottery.created_at.desc()).limit(5).all()
    recent_groups = db.query(PotteryGroup).order_by(PotteryGroup.created_at.desc()).limit(5).all()
    recent_tasks = db.query(RepairTask).order_by(RepairTask.created_at.desc()).limit(5).all()
    recent_repair_plans = db.query(RepairPlan).order_by(RepairPlan.created_at.desc()).limit(5).all()
    recent_logs = db.query(OperationLog).order_by(OperationLog.created_at.desc()).limit(10).all()
    recent_underwater_images = db.query(UnderwaterImage).order_by(UnderwaterImage.created_at.desc()).limit(5).all()

    damage_stats = {
        "轻微": db.query(Pottery).filter(Pottery.damage_level == "轻微").count(),
        "中度": db.query(Pottery).filter(Pottery.damage_level == "中度").count(),
        "严重": db.query(Pottery).filter(Pottery.damage_level == "严重").count()
    }

    repair_plan_stats = {
        "total": total_repair_plans,
        "pending_reviews": pending_reviews,
        "in_repair": in_repair_plans,
        "completed": completed_repair_plans
    }

    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "current_user": current_user,
        "total_potteries": total_potteries,
        "total_cleaned": total_cleaned,
        "total_groups": total_groups,
        "total_stored": total_stored,
        "total_tasks": total_tasks,
        "pending_tasks": pending_tasks,
        "pending_approvals": pending_approvals,
        "recent_potteries": recent_potteries,
        "recent_groups": recent_groups,
        "recent_tasks": recent_tasks,
        "recent_repair_plans": recent_repair_plans,
        "recent_logs": recent_logs,
        "recent_underwater_images": recent_underwater_images,
        "damage_stats": damage_stats,
        "underwater_image_stats": underwater_image_stats,
        "repair_plan_stats": repair_plan_stats,
        "today": date.today().isoformat()
    })


@app.get("/potteries", response_class=HTMLResponse)
async def list_potteries(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    keyword: Optional[str] = None,
    water_area: Optional[str] = None,
    damage_level: Optional[str] = None,
    current_status: Optional[str] = None
):
    query = db.query(Pottery)

    if keyword:
        query = query.filter(or_(
            Pottery.pottery_number.contains(keyword),
            Pottery.decoration_description.contains(keyword),
            Pottery.material.contains(keyword)
        ))
    if water_area:
        query = query.filter(Pottery.water_area == water_area)
    if damage_level:
        query = query.filter(Pottery.damage_level == damage_level)
    if current_status:
        query = query.filter(Pottery.current_status == current_status)

    potteries = query.order_by(Pottery.created_at.desc()).all()

    water_areas = [r[0] for r in db.query(Pottery.water_area).distinct().all() if r[0]]

    return templates.TemplateResponse("potteries.html", {
        "request": request,
        "current_user": current_user,
        "potteries": potteries,
        "keyword": keyword,
        "water_area": water_area,
        "damage_level": damage_level,
        "current_status": current_status,
        "water_areas": water_areas
    })


@app.get("/potteries/add", response_class=HTMLResponse)
async def add_pottery_page(request: Request, current_user: User = Depends(get_current_user)):
    return templates.TemplateResponse("pottery_form.html", {
        "request": request,
        "current_user": current_user,
        "pottery": None,
        "today": date.today().isoformat()
    })


@app.post("/potteries/add")
async def add_pottery(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    pottery_number: str = Form(...),
    water_area: str = Form(...),
    trench_number: str = Form(...),
    material: str = Form(...),
    decoration_description: str = Form(...),
    damage_level: str = Form(...),
    current_status: str = Form(...),
    recovery_date: date = Form(...)
):
    existing = db.query(Pottery).filter(Pottery.pottery_number == pottery_number).first()
    if existing:
        return templates.TemplateResponse("pottery_form.html", {
            "request": request,
            "current_user": current_user,
            "pottery": None,
            "today": date.today().isoformat(),
            "error": "陶片编号已存在"
        })

    if recovery_date > date.today():
        return templates.TemplateResponse("pottery_form.html", {
            "request": request,
            "current_user": current_user,
            "pottery": None,
            "today": date.today().isoformat(),
            "error": "出水日期不能晚于当前日期"
        })

    pottery = Pottery(
        pottery_number=pottery_number,
        water_area=water_area,
        trench_number=trench_number,
        material=material,
        decoration_description=decoration_description,
        damage_level=damage_level,
        current_status=current_status,
        recovery_date=recovery_date,
        is_locked=False
    )
    db.add(pottery)
    db.commit()
    db.refresh(pottery)

    log_operation(db, current_user, OperationType.CREATE, "pottery", pottery.id, f"创建陶片档案: {pottery_number}")

    return RedirectResponse(url="/potteries", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/potteries/{pottery_id}/edit", response_class=HTMLResponse)
async def edit_pottery_page(
    pottery_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    pottery = db.query(Pottery).filter(Pottery.id == pottery_id).first()
    if not pottery:
        return RedirectResponse(url="/potteries", status_code=status.HTTP_303_SEE_OTHER)
    
    if pottery.is_locked:
        return templates.TemplateResponse("pottery_detail.html", {
            "request": request,
            "current_user": current_user,
            "pottery": pottery,
            "cleaning_records": pottery.cleaning_records,
            "groups": pottery.pottery_groups,
            "storage_record": pottery.storage_record,
            "error": "该陶片已锁定，无法修改"
        })
    
    return templates.TemplateResponse("pottery_form.html", {
        "request": request,
        "current_user": current_user,
        "pottery": pottery,
        "today": date.today().isoformat()
    })


@app.post("/potteries/{pottery_id}/edit")
async def edit_pottery(
    pottery_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    water_area: str = Form(...),
    trench_number: str = Form(...),
    material: str = Form(...),
    decoration_description: str = Form(...),
    damage_level: str = Form(...),
    current_status: str = Form(...),
    recovery_date: date = Form(...)
):
    pottery = db.query(Pottery).filter(Pottery.id == pottery_id).first()
    if not pottery:
        return RedirectResponse(url="/potteries", status_code=status.HTTP_303_SEE_OTHER)
    
    if pottery.is_locked:
        return templates.TemplateResponse("pottery_detail.html", {
            "request": request,
            "current_user": current_user,
            "pottery": pottery,
            "cleaning_records": pottery.cleaning_records,
            "groups": pottery.pottery_groups,
            "storage_record": pottery.storage_record,
            "error": "该陶片已锁定，无法修改"
        })

    if recovery_date > date.today():
        return templates.TemplateResponse("pottery_form.html", {
            "request": request,
            "current_user": current_user,
            "pottery": pottery,
            "today": date.today().isoformat(),
            "error": "出水日期不能晚于当前日期"
        })

    pottery.water_area = water_area
    pottery.trench_number = trench_number
    pottery.material = material
    pottery.decoration_description = decoration_description
    pottery.damage_level = damage_level
    pottery.current_status = current_status
    pottery.recovery_date = recovery_date

    db.commit()
    log_operation(db, current_user, OperationType.UPDATE, "pottery", pottery_id, f"更新陶片档案: {pottery.pottery_number}")

    return RedirectResponse(url=f"/potteries/{pottery_id}", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/potteries/{pottery_id}", response_class=HTMLResponse)
async def pottery_detail(
    pottery_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    pottery = db.query(Pottery).filter(Pottery.id == pottery_id).first()
    if not pottery:
        return RedirectResponse(url="/potteries", status_code=status.HTTP_303_SEE_OTHER)

    cleaning_records = db.query(CleaningRecord).filter(CleaningRecord.pottery_id == pottery_id).order_by(CleaningRecord.cleaning_date.desc()).all()

    groups = db.query(PotteryGroup).join(PotteryGroupMember).filter(PotteryGroupMember.pottery_id == pottery_id).all()

    storage_record = db.query(StorageRecord).filter(StorageRecord.pottery_id == pottery_id).first()

    repair_tasks = db.query(RepairTask).filter(RepairTask.pottery_id == pottery_id).order_by(RepairTask.created_at.desc()).all()

    repair_plans = db.query(RepairPlan).filter(RepairPlan.pottery_id == pottery_id).order_by(RepairPlan.created_at.desc()).all()

    return templates.TemplateResponse("pottery_detail.html", {
        "request": request,
        "current_user": current_user,
        "pottery": pottery,
        "cleaning_records": cleaning_records,
        "groups": groups,
        "storage_record": storage_record,
        "repair_tasks": repair_tasks,
        "repair_plans": repair_plans
    })


@app.post("/potteries/{pottery_id}/images/upload")
async def upload_pottery_image(
    pottery_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    image: UploadFile = File(...),
    image_type: str = Form(...),
    description: str = Form("")
):
    pottery = db.query(Pottery).filter(Pottery.id == pottery_id).first()
    if not pottery:
        return RedirectResponse(url="/potteries", status_code=status.HTTP_303_SEE_OTHER)

    file_ext = os.path.splitext(image.filename)[1]
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(UPLOAD_DIR, unique_filename)

    with open(file_path, "wb") as buffer:
        content = await image.read()
        buffer.write(content)

    pottery_image = PotteryImage(
        pottery_id=pottery_id,
        image_path=f"/{file_path}",
        image_type=image_type,
        description=description,
        uploaded_by=current_user.id
    )
    db.add(pottery_image)
    db.commit()

    log_operation(db, current_user, OperationType.UPLOAD, "pottery_image", pottery_image.id, f"上传陶片图片: {pottery.pottery_number}")

    return RedirectResponse(url=f"/potteries/{pottery_id}", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/potteries/{pottery_id}/images/{image_id}/delete")
async def delete_pottery_image(
    pottery_id: int,
    image_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    image = db.query(PotteryImage).filter(PotteryImage.id == image_id, PotteryImage.pottery_id == pottery_id).first()
    if image:
        if os.path.exists(image.image_path.lstrip("/")):
            os.remove(image.image_path.lstrip("/"))
        db.delete(image)
        db.commit()
        log_operation(db, current_user, OperationType.DELETE, "pottery_image", image_id, f"删除陶片图片")

    return RedirectResponse(url=f"/potteries/{pottery_id}", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/potteries/{pottery_id}/cleaning/add", response_class=HTMLResponse)
async def add_cleaning_page(
    pottery_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    pottery = db.query(Pottery).filter(Pottery.id == pottery_id).first()
    if not pottery:
        return RedirectResponse(url="/potteries", status_code=status.HTTP_303_SEE_OTHER)
    return templates.TemplateResponse("cleaning_form.html", {
        "request": request,
        "current_user": current_user,
        "pottery": pottery,
        "today": date.today().isoformat()
    })


@app.post("/potteries/{pottery_id}/cleaning/add")
async def add_cleaning(
    pottery_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    cleaning_date: date = Form(...),
    cleaner: str = Form(...),
    cleaning_method: str = Form(...),
    cleaning_result: str = Form(...),
    notes: str = Form("")
):
    pottery = db.query(Pottery).filter(Pottery.id == pottery_id).first()
    if not pottery:
        return RedirectResponse(url="/potteries", status_code=status.HTTP_303_SEE_OTHER)

    if cleaning_date > date.today():
        return templates.TemplateResponse("cleaning_form.html", {
            "request": request,
            "current_user": current_user,
            "pottery": pottery,
            "today": date.today().isoformat(),
            "error": "清洗日期不能晚于当前日期"
        })

    cleaning = CleaningRecord(
        pottery_id=pottery_id,
        cleaning_date=cleaning_date,
        cleaner=cleaner,
        cleaning_method=cleaning_method,
        cleaning_result=cleaning_result,
        notes=notes
    )
    db.add(cleaning)
    db.commit()
    log_operation(db, current_user, OperationType.CREATE, "cleaning_record", cleaning.id, f"添加清洗记录: {pottery.pottery_number}")

    return RedirectResponse(url=f"/potteries/{pottery_id}", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/tasks", response_class=HTMLResponse)
async def list_tasks(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    status_filter: Optional[str] = None,
    priority_filter: Optional[str] = None,
    assignee_filter: Optional[int] = None
):
    query = db.query(RepairTask)

    if status_filter:
        query = query.filter(RepairTask.status == status_filter)
    if priority_filter:
        query = query.filter(RepairTask.priority == priority_filter)
    if assignee_filter:
        query = query.filter(RepairTask.assignee_id == assignee_filter)

    tasks = query.order_by(RepairTask.created_at.desc()).all()
    users = db.query(User).all()

    return templates.TemplateResponse("tasks.html", {
        "request": request,
        "current_user": current_user,
        "tasks": tasks,
        "users": users,
        "status_filter": status_filter,
        "priority_filter": priority_filter,
        "assignee_filter": assignee_filter
    })


@app.get("/tasks/add", response_class=HTMLResponse)
async def add_task_page(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    potteries = db.query(Pottery).all()
    users = db.query(User).all()
    return templates.TemplateResponse("task_form.html", {
        "request": request,
        "current_user": current_user,
        "task": None,
        "potteries": potteries,
        "users": users,
        "today": date.today().isoformat()
    })


@app.post("/tasks/add")
async def add_task(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    pottery_id: int = Form(...),
    title: str = Form(...),
    description: str = Form(""),
    priority: str = Form("普通"),
    assignee_id: Optional[int] = Form(None),
    due_date: Optional[date] = Form(None)
):
    task_number = f"TASK{datetime.now().strftime('%Y%m%d%H%M%S')}"

    task = RepairTask(
        task_number=task_number,
        pottery_id=pottery_id,
        title=title,
        description=description,
        priority=priority,
        status=TaskStatus.PENDING,
        creator_id=current_user.id,
        assignee_id=assignee_id,
        due_date=due_date
    )
    db.add(task)
    db.commit()
    db.refresh(task)

    log_operation(db, current_user, OperationType.CREATE, "repair_task", task.id, f"创建修复任务: {title}")

    return RedirectResponse(url="/tasks", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/tasks/{task_id}", response_class=HTMLResponse)
async def task_detail(
    task_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    task = db.query(RepairTask).filter(RepairTask.id == task_id).first()
    if not task:
        return RedirectResponse(url="/tasks", status_code=status.HTTP_303_SEE_OTHER)

    users = db.query(User).all()
    return templates.TemplateResponse("task_detail.html", {
        "request": request,
        "current_user": current_user,
        "task": task,
        "users": users
    })


@app.post("/tasks/{task_id}/edit")
async def edit_task(
    task_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    title: str = Form(...),
    description: str = Form(""),
    priority: str = Form(...),
    status: str = Form(...),
    assignee_id: Optional[int] = Form(None),
    due_date: Optional[date] = Form(None),
    result: str = Form("")
):
    task = db.query(RepairTask).filter(RepairTask.id == task_id).first()
    if not task:
        return RedirectResponse(url="/tasks", status_code=status.HTTP_303_SEE_OTHER)

    old_assignee = task.assignee_id
    old_status = task.status

    task.title = title
    task.description = description
    task.priority = priority
    task.status = status
    task.assignee_id = assignee_id
    task.due_date = due_date

    if status == TaskStatus.COMPLETED and old_status != TaskStatus.COMPLETED:
        task.completed_date = date.today()
        task.result = result
    elif status != TaskStatus.COMPLETED:
        task.completed_date = None
        task.result = None

    db.commit()

    if old_assignee != assignee_id:
        assignee = db.query(User).filter(User.id == assignee_id).first()
        log_operation(db, current_user, OperationType.ASSIGN, "repair_task", task_id, f"分派任务给: {assignee.full_name if assignee else '未指定'}")
    else:
        log_operation(db, current_user, OperationType.UPDATE, "repair_task", task_id, f"更新任务: {title}")

    return RedirectResponse(url=f"/tasks/{task_id}", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/groups", response_class=HTMLResponse)
async def list_groups(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    is_completed: Optional[str] = None
):
    query = db.query(PotteryGroup)
    if is_completed is not None and is_completed != "":
        query = query.filter(PotteryGroup.is_completed == (is_completed == "true"))

    groups = query.order_by(PotteryGroup.created_at.desc()).all()
    return templates.TemplateResponse("groups.html", {
        "request": request,
        "current_user": current_user,
        "groups": groups,
        "is_completed": is_completed
    })


@app.get("/groups/add", response_class=HTMLResponse)
async def add_group_page(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    stored_pottery_ids = [r[0] for r in db.query(StorageRecord.pottery_id).filter(StorageRecord.is_official == True).all()]
    active_group_members = db.query(PotteryGroupMember.pottery_id).join(PotteryGroup).filter(PotteryGroup.is_completed == False).all()
    active_pottery_ids = [r[0] for r in active_group_members]
    excluded_ids = stored_pottery_ids + active_pottery_ids

    available_potteries = db.query(Pottery).filter(~Pottery.id.in_(excluded_ids)).all()

    return templates.TemplateResponse("group_form.html", {
        "request": request,
        "current_user": current_user,
        "group": None,
        "available_potteries": available_potteries,
        "selected_potteries": []
    })


@app.post("/groups/add")
async def add_group(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    group_number: str = Form(...),
    confidence: int = Form(...),
    organizer: str = Form(...),
    notes: str = Form(""),
    change_description: str = Form("初始版本"),
    pottery_ids: List[int] = Form([])
):
    existing = db.query(PotteryGroup).filter(PotteryGroup.group_number == group_number).first()
    if existing:
        stored_pottery_ids = [r[0] for r in db.query(StorageRecord.pottery_id).filter(StorageRecord.is_official == True).all()]
        active_group_members = db.query(PotteryGroupMember.pottery_id).join(PotteryGroup).filter(PotteryGroup.is_completed == False).all()
        active_pottery_ids = [r[0] for r in active_group_members]
        excluded_ids = stored_pottery_ids + active_pottery_ids
        available_potteries = db.query(Pottery).filter(~Pottery.id.in_(excluded_ids)).all()

        return templates.TemplateResponse("group_form.html", {
            "request": request,
            "current_user": current_user,
            "group": None,
            "available_potteries": available_potteries,
            "selected_potteries": pottery_ids,
            "error": "组号已存在"
        })

    if confidence < 0 or confidence > 100:
        stored_pottery_ids = [r[0] for r in db.query(StorageRecord.pottery_id).filter(StorageRecord.is_official == True).all()]
        active_group_members = db.query(PotteryGroupMember.pottery_id).join(PotteryGroup).filter(PotteryGroup.is_completed == False).all()
        active_pottery_ids = [r[0] for r in active_group_members]
        excluded_ids = stored_pottery_ids + active_pottery_ids
        available_potteries = db.query(Pottery).filter(~Pottery.id.in_(excluded_ids)).all()

        return templates.TemplateResponse("group_form.html", {
            "request": request,
            "current_user": current_user,
            "group": None,
            "available_potteries": available_potteries,
            "selected_potteries": pottery_ids,
            "error": "拼合可信度范围为0-100"
        })

    for pid in pottery_ids:
        official_storage = db.query(StorageRecord).filter(
            StorageRecord.pottery_id == pid,
            StorageRecord.is_official == True
        ).first()
        if official_storage:
            pottery = db.query(Pottery).filter(Pottery.id == pid).first()
            stored_pottery_ids = [r[0] for r in db.query(StorageRecord.pottery_id).filter(StorageRecord.is_official == True).all()]
            active_group_members = db.query(PotteryGroupMember.pottery_id).join(PotteryGroup).filter(PotteryGroup.is_completed == False).all()
            active_pottery_ids = [r[0] for r in active_group_members]
            excluded_ids = stored_pottery_ids + active_pottery_ids
            available_potteries = db.query(Pottery).filter(~Pottery.id.in_(excluded_ids)).all()
            
            return templates.TemplateResponse("group_form.html", {
                "request": request,
                "current_user": current_user,
                "group": None,
                "available_potteries": available_potteries,
                "selected_potteries": [p for p in pottery_ids if p != pid],
                "error": f"陶片 {pottery.pottery_number if pottery else pid} 已正式入库，无法加入拼合组"
            })

        in_other_group = db.query(PotteryGroupMember).join(PotteryGroup).filter(
            PotteryGroupMember.pottery_id == pid,
            PotteryGroup.is_completed == False
        ).first()
        if in_other_group:
            pottery = db.query(Pottery).filter(Pottery.id == pid).first()
            stored_pottery_ids = [r[0] for r in db.query(StorageRecord.pottery_id).filter(StorageRecord.is_official == True).all()]
            active_group_members = db.query(PotteryGroupMember.pottery_id).join(PotteryGroup).filter(PotteryGroup.is_completed == False).all()
            active_pottery_ids = [r[0] for r in active_group_members]
            excluded_ids = stored_pottery_ids + active_pottery_ids
            available_potteries = db.query(Pottery).filter(~Pottery.id.in_(excluded_ids)).all()
            
            return templates.TemplateResponse("group_form.html", {
                "request": request,
                "current_user": current_user,
                "group": None,
                "available_potteries": available_potteries,
                "selected_potteries": [p for p in pottery_ids if p != pid],
                "error": f"陶片 {pottery.pottery_number if pottery else pid} 已在其他进行中的拼合组中"
            })

    group = PotteryGroup(
        group_number=group_number,
        confidence=confidence,
        organizer=organizer,
        notes=notes,
        is_completed=False,
        current_version=1
    )
    db.add(group)
    db.flush()

    for pid in pottery_ids:
        member = PotteryGroupMember(group_id=group.id, pottery_id=pid)
        db.add(member)

    version = GroupVersion(
        group_id=group.id,
        version_number=1,
        confidence=confidence,
        organizer=organizer,
        notes=notes,
        pottery_ids=",".join(map(str, pottery_ids)),
        change_description=change_description or "初始版本",
        created_by=current_user.id
    )
    db.add(version)

    db.commit()
    log_operation(db, current_user, OperationType.CREATE, "pottery_group", group.id, f"创建拼合组: {group_number}")

    return RedirectResponse(url="/groups", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/groups/{group_id}/edit", response_class=HTMLResponse)
async def edit_group_page(
    group_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    group = db.query(PotteryGroup).filter(PotteryGroup.id == group_id).first()
    if not group:
        return RedirectResponse(url="/groups", status_code=status.HTTP_303_SEE_OTHER)

    has_official = any(m.pottery.storage_record and m.pottery.storage_record.is_official for m in group.members)
    if has_official:
        groups = db.query(PotteryGroup).order_by(PotteryGroup.created_at.desc()).all()
        return templates.TemplateResponse("groups.html", {
            "request": request,
            "current_user": current_user,
            "groups": groups,
            "is_completed": None,
            "error": "包含已正式入库的陶片，无法修改"
        })

    stored_pottery_ids = [r[0] for r in db.query(StorageRecord.pottery_id).filter(StorageRecord.is_official == True).all()]
    active_group_members = db.query(PotteryGroupMember.pottery_id).join(PotteryGroup).filter(
        PotteryGroup.is_completed == False,
        PotteryGroup.id != group_id
    ).all()
    active_pottery_ids = [r[0] for r in active_group_members]
    excluded_ids = stored_pottery_ids + active_pottery_ids

    available_potteries = db.query(Pottery).filter(~Pottery.id.in_(excluded_ids)).all()
    selected_potteries = [m.pottery_id for m in group.members]

    versions = db.query(GroupVersion).filter(GroupVersion.group_id == group_id).order_by(GroupVersion.version_number.desc()).all()

    return templates.TemplateResponse("group_form.html", {
        "request": request,
        "current_user": current_user,
        "group": group,
        "available_potteries": available_potteries,
        "selected_potteries": selected_potteries,
        "versions": versions
    })


@app.post("/groups/{group_id}/edit")
async def edit_group(
    group_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    confidence: int = Form(...),
    organizer: str = Form(...),
    notes: str = Form(""),
    is_completed: str = Form("false"),
    change_description: str = Form(""),
    pottery_ids: List[int] = Form([])
):
    group = db.query(PotteryGroup).filter(PotteryGroup.id == group_id).first()
    if not group:
        return RedirectResponse(url="/groups", status_code=status.HTTP_303_SEE_OTHER)

    has_official = any(m.pottery.storage_record and m.pottery.storage_record.is_official for m in group.members)
    
    current_member_ids = {m.pottery_id for m in group.members}
    new_member_ids = set(pottery_ids)
    
    if has_official and current_member_ids != new_member_ids:
        groups = db.query(PotteryGroup).order_by(PotteryGroup.created_at.desc()).all()
        return templates.TemplateResponse("groups.html", {
            "request": request,
            "current_user": current_user,
            "groups": groups,
            "is_completed": None,
            "error": "该拼合组包含已正式入库的陶片，无法修改组成员"
        })

    if confidence < 0 or confidence > 100:
        stored_pottery_ids = [r[0] for r in db.query(StorageRecord.pottery_id).filter(StorageRecord.is_official == True).all()]
        active_group_members = db.query(PotteryGroupMember.pottery_id).join(PotteryGroup).filter(
            PotteryGroup.is_completed == False,
            PotteryGroup.id != group_id
        ).all()
        active_pottery_ids = [r[0] for r in active_group_members]
        excluded_ids = stored_pottery_ids + active_pottery_ids
        available_potteries = db.query(Pottery).filter(~Pottery.id.in_(excluded_ids)).all()

        return templates.TemplateResponse("group_form.html", {
            "request": request,
            "current_user": current_user,
            "group": group,
            "available_potteries": available_potteries,
            "selected_potteries": pottery_ids,
            "error": "拼合可信度范围为0-100"
        })

    if not has_official:
        for pid in pottery_ids:
            if pid not in current_member_ids:
                official_storage = db.query(StorageRecord).filter(
                    StorageRecord.pottery_id == pid,
                    StorageRecord.is_official == True
                ).first()
                if official_storage:
                    pottery = db.query(Pottery).filter(Pottery.id == pid).first()
                    stored_pottery_ids = [r[0] for r in db.query(StorageRecord.pottery_id).filter(StorageRecord.is_official == True).all()]
                    active_group_members = db.query(PotteryGroupMember.pottery_id).join(PotteryGroup).filter(
                        PotteryGroup.is_completed == False,
                        PotteryGroup.id != group_id
                    ).all()
                    active_pottery_ids = [r[0] for r in active_group_members]
                    excluded_ids = stored_pottery_ids + active_pottery_ids
                    available_potteries = db.query(Pottery).filter(~Pottery.id.in_(excluded_ids)).all()
                    
                    return templates.TemplateResponse("group_form.html", {
                        "request": request,
                        "current_user": current_user,
                        "group": group,
                        "available_potteries": available_potteries,
                        "selected_potteries": [p for p in pottery_ids if p != pid],
                        "error": f"陶片 {pottery.pottery_number if pottery else pid} 已正式入库，无法加入拼合组"
                    })

                in_other_group = db.query(PotteryGroupMember).join(PotteryGroup).filter(
                    PotteryGroupMember.pottery_id == pid,
                    PotteryGroup.is_completed == False,
                    PotteryGroup.id != group_id
                ).first()
                if in_other_group:
                    pottery = db.query(Pottery).filter(Pottery.id == pid).first()
                    stored_pottery_ids = [r[0] for r in db.query(StorageRecord.pottery_id).filter(StorageRecord.is_official == True).all()]
                    active_group_members = db.query(PotteryGroupMember.pottery_id).join(PotteryGroup).filter(
                        PotteryGroup.is_completed == False,
                        PotteryGroup.id != group_id
                    ).all()
                    active_pottery_ids = [r[0] for r in active_group_members]
                    excluded_ids = stored_pottery_ids + active_pottery_ids
                    available_potteries = db.query(Pottery).filter(~Pottery.id.in_(excluded_ids)).all()
                    
                    return templates.TemplateResponse("group_form.html", {
                        "request": request,
                        "current_user": current_user,
                        "group": group,
                        "available_potteries": available_potteries,
                        "selected_potteries": [p for p in pottery_ids if p != pid],
                        "error": f"陶片 {pottery.pottery_number if pottery else pid} 已在其他进行中的拼合组中"
                    })

    members_changed = current_member_ids != new_member_ids
    metadata_changed = (group.confidence != confidence or 
                        group.organizer != organizer or 
                        group.notes != notes)

    group.confidence = confidence
    group.organizer = organizer
    group.notes = notes
    group.is_completed = is_completed == "true"

    if not has_official:
        db.query(PotteryGroupMember).filter(PotteryGroupMember.group_id == group_id).delete()
        for pid in pottery_ids:
            member = PotteryGroupMember(group_id=group.id, pottery_id=pid)
            db.add(member)

    if members_changed or metadata_changed:
        new_version = group.current_version + 1
        group.current_version = new_version
        
        version = GroupVersion(
            group_id=group.id,
            version_number=new_version,
            confidence=confidence,
            organizer=organizer,
            notes=notes,
            pottery_ids=",".join(map(str, pottery_ids)),
            change_description=change_description or f"版本 {new_version} 更新",
            created_by=current_user.id
        )
        db.add(version)

    db.commit()
    log_operation(db, current_user, OperationType.UPDATE, "pottery_group", group_id, f"更新拼合组: {group.group_number}")

    return RedirectResponse(url=f"/groups/{group_id}", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/groups/{group_id}", response_class=HTMLResponse)
async def group_detail(
    group_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    group = db.query(PotteryGroup).filter(PotteryGroup.id == group_id).first()
    if not group:
        return RedirectResponse(url="/groups", status_code=status.HTTP_303_SEE_OTHER)

    severe_count = 0
    official_count = 0
    for m in group.members:
        if m.pottery.damage_level == "严重":
            severe_count += 1
        if m.pottery.storage_record and m.pottery.storage_record.is_official:
            official_count += 1

    versions = db.query(GroupVersion).filter(GroupVersion.group_id == group_id).order_by(GroupVersion.version_number.desc()).all()

    return templates.TemplateResponse("group_detail.html", {
        "request": request,
        "current_user": current_user,
        "group": group,
        "severe_count": severe_count,
        "official_count": official_count,
        "versions": versions
    })


@app.get("/storage", response_class=HTMLResponse)
async def list_storage(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    is_official: Optional[str] = None,
    approval_status: Optional[str] = None
):
    query = db.query(StorageRecord)
    if is_official is not None and is_official != "":
        query = query.filter(StorageRecord.is_official == (is_official == "true"))
    if approval_status:
        query = query.filter(StorageRecord.approval_status == approval_status)

    records = query.order_by(StorageRecord.storage_date.desc()).all()
    return templates.TemplateResponse("storage.html", {
        "request": request,
        "current_user": current_user,
        "records": records,
        "is_official": is_official,
        "approval_status": approval_status
    })


@app.get("/storage/add", response_class=HTMLResponse)
async def add_storage_page(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    stored_ids = [r[0] for r in db.query(StorageRecord.pottery_id).all()]
    available_potteries = db.query(Pottery).filter(~Pottery.id.in_(stored_ids)).all()

    return templates.TemplateResponse("storage_form.html", {
        "request": request,
        "current_user": current_user,
        "record": None,
        "available_potteries": available_potteries,
        "today": date.today().isoformat()
    })


@app.post("/storage/add")
async def add_storage(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    pottery_id: int = Form(...),
    storage_date: date = Form(...),
    location: str = Form(...),
    registrar: str = Form(...),
    notes: str = Form(""),
    is_official: str = Form("false")
):
    existing = db.query(StorageRecord).filter(StorageRecord.pottery_id == pottery_id).first()
    if existing:
        return RedirectResponse(url="/storage", status_code=status.HTTP_303_SEE_OTHER)

    if storage_date > date.today():
        stored_ids = [r[0] for r in db.query(StorageRecord.pottery_id).all()]
        available_potteries = db.query(Pottery).filter(~Pottery.id.in_(stored_ids)).all()
        return templates.TemplateResponse("storage_form.html", {
            "request": request,
            "current_user": current_user,
            "record": None,
            "available_potteries": available_potteries,
            "today": date.today().isoformat(),
            "error": "入库日期不能晚于当前日期"
        })

    record = StorageRecord(
        pottery_id=pottery_id,
        storage_date=storage_date,
        location=location,
        registrar=registrar,
        notes=notes,
        is_official=is_official == "true",
        approval_status=ApprovalStatus.DRAFT
    )
    db.add(record)
    db.commit()
    db.refresh(record)

    log_operation(db, current_user, OperationType.CREATE, "storage_record", record.id, f"创建立库记录")

    return RedirectResponse(url="/storage", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/storage/{record_id}/submit")
async def submit_for_approval(
    record_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    record = db.query(StorageRecord).filter(StorageRecord.id == record_id).first()
    if record:
        record.approval_status = ApprovalStatus.SUBMITTED
        db.commit()
        log_operation(db, current_user, OperationType.SUBMIT, "storage_record", record_id, f"提交入库审批")

    return RedirectResponse(url="/storage", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/storage/{record_id}/approve")
async def approve_storage(
    record_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    record = db.query(StorageRecord).filter(StorageRecord.id == record_id).first()
    if record:
        record.approval_status = ApprovalStatus.APPROVED
        record.is_official = True
        db.commit()

        approval = StorageApproval(
            storage_id=record_id,
            approver_id=current_user.id,
            approval_status=ApprovalStatus.APPROVED,
            comments="审批通过"
        )
        db.add(approval)
        db.commit()

        pottery = db.query(Pottery).filter(Pottery.id == record.pottery_id).first()
        if pottery:
            pottery.is_locked = True
            db.commit()

        log_operation(db, current_user, OperationType.APPROVE, "storage_record", record_id, f"审批通过入库记录")

    return RedirectResponse(url="/storage", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/storage/{record_id}/reject")
async def reject_storage(
    record_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    record = db.query(StorageRecord).filter(StorageRecord.id == record_id).first()
    if record:
        record.approval_status = ApprovalStatus.REJECTED
        db.commit()

        approval = StorageApproval(
            storage_id=record_id,
            approver_id=current_user.id,
            approval_status=ApprovalStatus.REJECTED,
            comments="审批驳回"
        )
        db.add(approval)
        db.commit()

        log_operation(db, current_user, OperationType.REJECT, "storage_record", record_id, f"审批驳回入库记录")

    return RedirectResponse(url="/storage", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/logs", response_class=HTMLResponse)
async def list_logs(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    operation_type: Optional[str] = None,
    target_type: Optional[str] = None,
    user_id: Optional[int] = None
):
    query = db.query(OperationLog)

    if operation_type:
        query = query.filter(OperationLog.operation_type == operation_type)
    if target_type:
        query = query.filter(OperationLog.target_type == target_type)
    if user_id:
        query = query.filter(OperationLog.user_id == user_id)

    logs = query.order_by(OperationLog.created_at.desc()).limit(200).all()
    users = db.query(User).all()

    return templates.TemplateResponse("logs.html", {
        "request": request,
        "current_user": current_user,
        "logs": logs,
        "users": users,
        "operation_type": operation_type,
        "target_type": target_type,
        "user_id": user_id
    })


@app.get("/reports", response_class=HTMLResponse)
async def reports_page(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    water_areas = [r[0] for r in db.query(Pottery.water_area).distinct().all() if r[0]]
    materials = [r[0] for r in db.query(Pottery.material).distinct().all() if r[0]]
    statuses = [r[0] for r in db.query(Pottery.current_status).distinct().all() if r[0]]

    return templates.TemplateResponse("reports.html", {
        "request": request,
        "current_user": current_user,
        "water_areas": water_areas,
        "materials": materials,
        "statuses": statuses,
        "excel_available": EXCEL_AVAILABLE
    })


@app.get("/reports/export")
async def export_report(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    water_area: Optional[str] = None,
    material: Optional[str] = None,
    current_status: Optional[str] = None,
    report_type: str = "potteries"
):
    if not EXCEL_AVAILABLE:
        return {"error": "Excel export not available. Please install openpyxl."}

    wb = openpyxl.Workbook()
    
    if report_type == "potteries":
        ws = wb.active
        ws.title = "陶片档案"
        
        headers = ["编号", "陶片编号", "水域", "探方号", "材质", "纹饰描述", "残损程度", "状态", "出水日期"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        query = db.query(Pottery)
        if water_area:
            query = query.filter(Pottery.water_area == water_area)
        if material:
            query = query.filter(Pottery.material == material)
        if current_status:
            query = query.filter(Pottery.current_status == current_status)
        
        potteries = query.order_by(Pottery.pottery_number).all()
        
        for idx, p in enumerate(potteries, 1):
            ws.append([
                idx, p.pottery_number, p.water_area, p.trench_number,
                p.material, p.decoration_description, p.damage_level,
                p.current_status, p.recovery_date.strftime("%Y-%m-%d") if p.recovery_date else ""
            ])

    elif report_type == "storage":
        ws = wb.active
        ws.title = "入库记录"
        
        headers = ["编号", "陶片编号", "入库日期", "存放位置", "登记人", "是否正式", "审批状态"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        query = db.query(StorageRecord).join(Pottery)
        records = query.order_by(StorageRecord.storage_date.desc()).all()
        
        for idx, r in enumerate(records, 1):
            ws.append([
                idx, r.pottery.pottery_number,
                r.storage_date.strftime("%Y-%m-%d") if r.storage_date else "",
                r.location, r.registrar,
                "是" if r.is_official else "否",
                r.approval_status
            ])

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{report_type}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/search", response_class=HTMLResponse)
async def search_page(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    q: Optional[str] = None,
    type: str = "all"
):
    results = {
        "potteries": [],
        "groups": [],
        "storage": [],
        "tasks": []
    }

    if q:
        if type in ["all", "pottery"]:
            results["potteries"] = db.query(Pottery).filter(or_(
                Pottery.pottery_number.contains(q),
                Pottery.water_area.contains(q),
                Pottery.decoration_description.contains(q),
                Pottery.material.contains(q)
            )).all()

        if type in ["all", "group"]:
            results["groups"] = db.query(PotteryGroup).filter(or_(
                PotteryGroup.group_number.contains(q),
                PotteryGroup.organizer.contains(q),
                PotteryGroup.notes.contains(q)
            )).all()

        if type in ["all", "storage"]:
            results["storage"] = db.query(StorageRecord).filter(or_(
                StorageRecord.location.contains(q),
                StorageRecord.registrar.contains(q)
            )).join(Pottery).filter(or_(
                Pottery.pottery_number.contains(q),
                StorageRecord.location.contains(q),
                StorageRecord.registrar.contains(q)
            )).all()

        if type in ["all", "task"]:
            results["tasks"] = db.query(RepairTask).filter(or_(
                RepairTask.task_number.contains(q),
                RepairTask.title.contains(q),
                RepairTask.description.contains(q)
            )).all()

    return templates.TemplateResponse("search.html", {
        "request": request,
        "current_user": current_user,
        "q": q,
        "type": type,
        "results": results
    })


@app.get("/underwater_images", response_class=HTMLResponse)
async def list_underwater_images(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    pottery_id: Optional[int] = None,
    group_id: Optional[int] = None,
    water_area: Optional[str] = None,
    image_type: Optional[str] = None,
    keyword: Optional[str] = None
):
    query = db.query(UnderwaterImage)
    
    if pottery_id:
        query = query.filter(UnderwaterImage.pottery_id == pottery_id)
    if group_id:
        query = query.filter(UnderwaterImage.group_id == group_id)
    if water_area:
        query = query.filter(UnderwaterImage.water_area == water_area)
    if image_type:
        query = query.filter(UnderwaterImage.image_type == image_type)
    if keyword:
        query = query.filter(or_(
            UnderwaterImage.image_number.contains(keyword),
            UnderwaterImage.description.contains(keyword),
            UnderwaterImage.photographer.contains(keyword)
        ))
    
    images = query.order_by(UnderwaterImage.created_at.desc()).all()
    
    water_areas = [r[0] for r in db.query(UnderwaterImage.water_area).distinct().all() if r[0]]
    potteries = db.query(Pottery).all()
    groups = db.query(PotteryGroup).all()
    
    return templates.TemplateResponse("underwater_images.html", {
        "request": request,
        "current_user": current_user,
        "images": images,
        "pottery_id": pottery_id,
        "group_id": group_id,
        "water_area": water_area,
        "image_type": image_type,
        "keyword": keyword,
        "water_areas": water_areas,
        "potteries": potteries,
        "groups": groups
    })


@app.get("/underwater_images/add", response_class=HTMLResponse)
async def add_underwater_image_page(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    pottery_id: Optional[int] = None
):
    potteries = db.query(Pottery).all()
    groups = db.query(PotteryGroup).all()
    selected_pottery = None
    if pottery_id:
        selected_pottery = db.query(Pottery).filter(Pottery.id == pottery_id).first()
    
    return templates.TemplateResponse("underwater_image_form.html", {
        "request": request,
        "current_user": current_user,
        "image": None,
        "potteries": potteries,
        "groups": groups,
        "selected_pottery": selected_pottery,
        "today": datetime.now().strftime("%Y-%m-%dT%H:%M")
    })


@app.post("/underwater_images/add")
async def add_underwater_image(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    pottery_id: int = Form(...),
    group_id: Optional[int] = Form(None),
    image_number: str = Form(...),
    image_type: str = Form(...),
    file_type: str = Form("image"),
    description: str = Form(""),
    coordinate_x: str = Form(""),
    coordinate_y: str = Form(""),
    coordinate_z: str = Form(""),
    depth: str = Form(""),
    shooting_time: Optional[datetime] = Form(None),
    photographer: str = Form(""),
    trench_number: str = Form(""),
    water_area: str = Form(""),
    file: UploadFile = File(...)
):
    existing = db.query(UnderwaterImage).filter(UnderwaterImage.image_number == image_number).first()
    if existing:
        potteries = db.query(Pottery).all()
        groups = db.query(PotteryGroup).all()
        return templates.TemplateResponse("underwater_image_form.html", {
            "request": request,
            "current_user": current_user,
            "image": None,
            "potteries": potteries,
            "groups": groups,
            "selected_pottery": None,
            "today": datetime.now().strftime("%Y-%m-%dT%H:%M"),
            "error": "影像编号已存在"
        })
    
    file_ext = os.path.splitext(file.filename)[1]
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(UPLOAD_DIR, unique_filename)
    
    with open(file_path, "wb") as buffer:
        content = await file.read()
        buffer.write(content)
    
    pottery = db.query(Pottery).filter(Pottery.id == pottery_id).first()
    if pottery and not water_area:
        water_area = pottery.water_area
    if pottery and not trench_number:
        trench_number = pottery.trench_number
    
    image = UnderwaterImage(
        pottery_id=pottery_id,
        group_id=group_id,
        image_number=image_number,
        image_type=image_type,
        file_type=file_type,
        file_path=f"/{file_path}",
        description=description,
        coordinate_x=coordinate_x,
        coordinate_y=coordinate_y,
        coordinate_z=coordinate_z,
        depth=depth,
        shooting_time=shooting_time,
        photographer=photographer,
        trench_number=trench_number,
        water_area=water_area,
        uploaded_by=current_user.id
    )
    db.add(image)
    db.commit()
    db.refresh(image)
    
    log_operation(db, current_user, OperationType.UPLOAD, "underwater_image", image.id, f"上传水下影像: {image_number}")
    
    return RedirectResponse(url="/underwater_images", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/underwater_images/{image_id}", response_class=HTMLResponse)
async def underwater_image_detail(
    image_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    image = db.query(UnderwaterImage).filter(UnderwaterImage.id == image_id).first()
    if not image:
        return RedirectResponse(url="/underwater_images", status_code=status.HTTP_303_SEE_OTHER)
    
    pottery_images = db.query(UnderwaterImage).filter(
        UnderwaterImage.pottery_id == image.pottery_id
    ).order_by(UnderwaterImage.shooting_time.asc()).all()
    
    timeline_data = []
    for img in pottery_images:
        timeline_data.append({
            "id": img.id,
            "image_number": img.image_number,
            "shooting_time": img.shooting_time,
            "image_type": img.image_type,
            "file_path": img.file_path,
            "description": img.description,
            "is_current": img.id == image_id
        })
    
    return templates.TemplateResponse("underwater_image_detail.html", {
        "request": request,
        "current_user": current_user,
        "image": image,
        "timeline_data": timeline_data
    })


@app.get("/underwater_images/{image_id}/edit", response_class=HTMLResponse)
async def edit_underwater_image_page(
    image_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    image = db.query(UnderwaterImage).filter(UnderwaterImage.id == image_id).first()
    if not image:
        return RedirectResponse(url="/underwater_images", status_code=status.HTTP_303_SEE_OTHER)
    
    potteries = db.query(Pottery).all()
    groups = db.query(PotteryGroup).all()
    
    return templates.TemplateResponse("underwater_image_form.html", {
        "request": request,
        "current_user": current_user,
        "image": image,
        "potteries": potteries,
        "groups": groups,
        "selected_pottery": None,
        "today": datetime.now().strftime("%Y-%m-%dT%H:%M")
    })


@app.post("/underwater_images/{image_id}/edit")
async def edit_underwater_image(
    image_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    group_id: Optional[int] = Form(None),
    image_type: str = Form(...),
    description: str = Form(""),
    coordinate_x: str = Form(""),
    coordinate_y: str = Form(""),
    coordinate_z: str = Form(""),
    depth: str = Form(""),
    shooting_time: Optional[datetime] = Form(None),
    photographer: str = Form(""),
    trench_number: str = Form(""),
    water_area: str = Form("")
):
    image = db.query(UnderwaterImage).filter(UnderwaterImage.id == image_id).first()
    if not image:
        return RedirectResponse(url="/underwater_images", status_code=status.HTTP_303_SEE_OTHER)
    
    image.group_id = group_id
    image.image_type = image_type
    image.description = description
    image.coordinate_x = coordinate_x
    image.coordinate_y = coordinate_y
    image.coordinate_z = coordinate_z
    image.depth = depth
    image.shooting_time = shooting_time
    image.photographer = photographer
    image.trench_number = trench_number
    image.water_area = water_area
    
    db.commit()
    log_operation(db, current_user, OperationType.UPDATE, "underwater_image", image_id, f"更新水下影像: {image.image_number}")
    
    return RedirectResponse(url=f"/underwater_images/{image_id}", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/underwater_images/{image_id}/delete")
async def delete_underwater_image(
    image_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    image = db.query(UnderwaterImage).filter(UnderwaterImage.id == image_id).first()
    if image:
        if os.path.exists(image.file_path.lstrip("/")):
            os.remove(image.file_path.lstrip("/"))
        db.delete(image)
        db.commit()
        log_operation(db, current_user, OperationType.DELETE, "underwater_image", image_id, f"删除水下影像: {image.image_number}")
    
    return RedirectResponse(url="/underwater_images", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/repair_plans", response_class=HTMLResponse)
async def list_repair_plans(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    status_filter: Optional[str] = None,
    progress_filter: Optional[str] = None,
    restorer_filter: Optional[int] = None,
    pottery_id: Optional[int] = None
):
    query = db.query(RepairPlan)

    if status_filter:
        query = query.filter(RepairPlan.status == status_filter)
    if progress_filter:
        query = query.filter(RepairPlan.progress == progress_filter)
    if restorer_filter:
        query = query.filter(RepairPlan.restorer_id == restorer_filter)
    if pottery_id:
        query = query.filter(RepairPlan.pottery_id == pottery_id)

    plans = query.order_by(RepairPlan.created_at.desc()).all()
    users = db.query(User).all()
    potteries = db.query(Pottery).all()

    return templates.TemplateResponse("repair_plans.html", {
        "request": request,
        "current_user": current_user,
        "plans": plans,
        "users": users,
        "potteries": potteries,
        "status_filter": status_filter,
        "progress_filter": progress_filter,
        "restorer_filter": restorer_filter,
        "pottery_id": pottery_id
    })


@app.get("/repair_plans/add", response_class=HTMLResponse)
async def add_repair_plan_page(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    pottery_id: Optional[int] = None,
    task_id: Optional[int] = None,
    group_id: Optional[int] = None
):
    available_potteries = db.query(Pottery).filter(Pottery.is_locked == False).all()
    users = db.query(User).all()
    tasks = db.query(RepairTask).filter(RepairTask.status != TaskStatus.COMPLETED).all()
    groups = db.query(PotteryGroup).filter(PotteryGroup.is_completed == False).all()

    selected_pottery = None
    selected_task = None
    selected_group = None

    if pottery_id:
        selected_pottery = db.query(Pottery).filter(Pottery.id == pottery_id).first()
    if task_id:
        selected_task = db.query(RepairTask).filter(RepairTask.id == task_id).first()
    if group_id:
        selected_group = db.query(PotteryGroup).filter(PotteryGroup.id == group_id).first()

    return templates.TemplateResponse("repair_plan_form.html", {
        "request": request,
        "current_user": current_user,
        "plan": None,
        "potteries": available_potteries,
        "users": users,
        "tasks": tasks,
        "groups": groups,
        "selected_pottery": selected_pottery,
        "selected_task": selected_task,
        "selected_group": selected_group,
        "today": date.today().isoformat(),
        "statuses": [e.value for e in RepairPlanStatus],
        "progresses": [e.value for e in RepairProgress]
    })


@app.post("/repair_plans/add")
async def add_repair_plan(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    pottery_id: int = Form(...),
    group_id: Optional[int] = Form(None),
    task_id: Optional[int] = Form(None),
    plan_name: str = Form(...),
    plan_description: str = Form(""),
    repair_method: str = Form(""),
    materials_used: str = Form(""),
    estimated_duration: Optional[int] = Form(None),
    expected_completion_date: Optional[date] = Form(None),
    restorer_id: Optional[int] = Form(None),
    restorer_name: str = Form("")
):
    pottery = db.query(Pottery).filter(Pottery.id == pottery_id).first()
    if not pottery:
        return RedirectResponse(url="/repair_plans", status_code=status.HTTP_303_SEE_OTHER)

    if pottery.is_locked:
        available_potteries = db.query(Pottery).filter(Pottery.is_locked == False).all()
        users = db.query(User).all()
        tasks = db.query(RepairTask).filter(RepairTask.status != TaskStatus.COMPLETED).all()
        groups = db.query(PotteryGroup).filter(PotteryGroup.is_completed == False).all()
        return templates.TemplateResponse("repair_plan_form.html", {
            "request": request,
            "current_user": current_user,
            "plan": None,
            "potteries": available_potteries,
            "users": users,
            "tasks": tasks,
            "groups": groups,
            "selected_pottery": None,
            "selected_task": None,
            "selected_group": None,
            "today": date.today().isoformat(),
            "statuses": [e.value for e in RepairPlanStatus],
            "progresses": [e.value for e in RepairProgress],
            "error": "该陶片已锁定，无法创建修复方案"
        })

    if expected_completion_date and expected_completion_date < date.today():
        available_potteries = db.query(Pottery).filter(Pottery.is_locked == False).all()
        users = db.query(User).all()
        tasks = db.query(RepairTask).filter(RepairTask.status != TaskStatus.COMPLETED).all()
        groups = db.query(PotteryGroup).filter(PotteryGroup.is_completed == False).all()
        return templates.TemplateResponse("repair_plan_form.html", {
            "request": request,
            "current_user": current_user,
            "plan": None,
            "potteries": available_potteries,
            "users": users,
            "tasks": tasks,
            "groups": groups,
            "selected_pottery": pottery,
            "selected_task": None,
            "selected_group": None,
            "today": date.today().isoformat(),
            "statuses": [e.value for e in RepairPlanStatus],
            "progresses": [e.value for e in RepairProgress],
            "error": "预计完成日期不能早于当前日期"
        })

    plan_number = f"PLAN{datetime.now().strftime('%Y%m%d%H%M%S')}"

    plan = RepairPlan(
        plan_number=plan_number,
        pottery_id=pottery_id,
        group_id=group_id,
        task_id=task_id,
        plan_name=plan_name,
        plan_description=plan_description,
        repair_method=repair_method,
        materials_used=materials_used,
        estimated_duration=estimated_duration,
        expected_completion_date=expected_completion_date,
        status=RepairPlanStatus.DRAFT,
        progress=RepairProgress.NOT_STARTED,
        restorer_id=restorer_id,
        restorer_name=restorer_name,
        created_by=current_user.id
    )
    db.add(plan)
    db.commit()
    db.refresh(plan)

    log_operation(db, current_user, OperationType.CREATE, "repair_plan", plan.id, f"创建修复方案: {plan_name}")

    return RedirectResponse(url="/repair_plans", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/repair_plans/{plan_id}", response_class=HTMLResponse)
async def repair_plan_detail(
    plan_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    plan = db.query(RepairPlan).filter(RepairPlan.id == plan_id).first()
    if not plan:
        return RedirectResponse(url="/repair_plans", status_code=status.HTTP_303_SEE_OTHER)

    users = db.query(User).all()
    review_records = db.query(ReviewRecord).filter(ReviewRecord.repair_plan_id == plan_id).order_by(ReviewRecord.created_at.desc()).all()
    comparison_images = db.query(RepairComparisonImage).filter(RepairComparisonImage.repair_plan_id == plan_id).order_by(RepairComparisonImage.created_at.asc()).all()

    return templates.TemplateResponse("repair_plan_detail.html", {
        "request": request,
        "current_user": current_user,
        "plan": plan,
        "users": users,
        "review_records": review_records,
        "comparison_images": comparison_images,
        "statuses": [e.value for e in RepairPlanStatus],
        "progresses": [e.value for e in RepairProgress],
        "today": date.today().isoformat()
    })


@app.get("/repair_plans/{plan_id}/edit", response_class=HTMLResponse)
async def edit_repair_plan_page(
    plan_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    plan = db.query(RepairPlan).filter(RepairPlan.id == plan_id).first()
    if not plan:
        return RedirectResponse(url="/repair_plans", status_code=status.HTTP_303_SEE_OTHER)

    if plan.status in [RepairPlanStatus.APPROVED, RepairPlanStatus.COMPLETED]:
        return templates.TemplateResponse("repair_plan_detail.html", {
            "request": request,
            "current_user": current_user,
            "plan": plan,
            "users": [],
            "review_records": [],
            "comparison_images": [],
            "statuses": [e.value for e in RepairPlanStatus],
            "progresses": [e.value for e in RepairProgress],
            "today": date.today().isoformat(),
            "error": "该方案已通过或已完成，无法修改"
        })

    available_potteries = db.query(Pottery).filter(Pottery.is_locked == False).all()
    users = db.query(User).all()
    tasks = db.query(RepairTask).filter(RepairTask.status != TaskStatus.COMPLETED).all()
    groups = db.query(PotteryGroup).filter(PotteryGroup.is_completed == False).all()

    return templates.TemplateResponse("repair_plan_form.html", {
        "request": request,
        "current_user": current_user,
        "plan": plan,
        "potteries": available_potteries,
        "users": users,
        "tasks": tasks,
        "groups": groups,
        "selected_pottery": None,
        "selected_task": None,
        "selected_group": None,
        "today": date.today().isoformat(),
        "statuses": [e.value for e in RepairPlanStatus],
        "progresses": [e.value for e in RepairProgress]
    })


@app.post("/repair_plans/{plan_id}/edit")
async def edit_repair_plan(
    plan_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    plan_name: str = Form(...),
    plan_description: str = Form(""),
    repair_method: str = Form(""),
    materials_used: str = Form(""),
    estimated_duration: Optional[int] = Form(None),
    expected_completion_date: Optional[date] = Form(None),
    restorer_id: Optional[int] = Form(None),
    restorer_name: str = Form("")
):
    plan = db.query(RepairPlan).filter(RepairPlan.id == plan_id).first()
    if not plan:
        return RedirectResponse(url="/repair_plans", status_code=status.HTTP_303_SEE_OTHER)

    if plan.status in [RepairPlanStatus.APPROVED, RepairPlanStatus.COMPLETED]:
        return templates.TemplateResponse("repair_plan_detail.html", {
            "request": request,
            "current_user": current_user,
            "plan": plan,
            "users": [],
            "review_records": [],
            "comparison_images": [],
            "statuses": [e.value for e in RepairPlanStatus],
            "progresses": [e.value for e in RepairProgress],
            "today": date.today().isoformat(),
            "error": "该方案已通过或已完成，无法修改"
        })

    if expected_completion_date and expected_completion_date < date.today():
        available_potteries = db.query(Pottery).filter(Pottery.is_locked == False).all()
        users = db.query(User).all()
        tasks = db.query(RepairTask).filter(RepairTask.status != TaskStatus.COMPLETED).all()
        groups = db.query(PotteryGroup).filter(PotteryGroup.is_completed == False).all()
        return templates.TemplateResponse("repair_plan_form.html", {
            "request": request,
            "current_user": current_user,
            "plan": plan,
            "potteries": available_potteries,
            "users": users,
            "tasks": tasks,
            "groups": groups,
            "selected_pottery": None,
            "selected_task": None,
            "selected_group": None,
            "today": date.today().isoformat(),
            "statuses": [e.value for e in RepairPlanStatus],
            "progresses": [e.value for e in RepairProgress],
            "error": "预计完成日期不能早于当前日期"
        })

    plan.plan_name = plan_name
    plan.plan_description = plan_description
    plan.repair_method = repair_method
    plan.materials_used = materials_used
    plan.estimated_duration = estimated_duration
    plan.expected_completion_date = expected_completion_date
    plan.restorer_id = restorer_id
    plan.restorer_name = restorer_name

    db.commit()
    log_operation(db, current_user, OperationType.UPDATE, "repair_plan", plan_id, f"更新修复方案: {plan_name}")

    return RedirectResponse(url=f"/repair_plans/{plan_id}", status_code=status.HTTP_303_SEE_OTHER)


@app.post("/repair_plans/{plan_id}/submit")
async def submit_repair_plan(
    plan_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    plan = db.query(RepairPlan).filter(RepairPlan.id == plan_id).first()
    if plan:
        plan.status = RepairPlanStatus.IN_REVIEW
        plan.submitted_at = datetime.now()
        db.commit()
        log_operation(db, current_user, OperationType.SUBMIT, "repair_plan", plan_id, f"提交修复方案复核: {plan.plan_name}")

    return RedirectResponse(url=f"/repair_plans/{plan_id}", status_code=status.HTTP_303_SEE_OTHER)


@app.post("/repair_plans/{plan_id}/progress")
async def update_repair_progress(
    plan_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    progress: str = Form(...),
    status: Optional[str] = Form(None)
):
    plan = db.query(RepairPlan).filter(RepairPlan.id == plan_id).first()
    if plan:
        old_progress = plan.progress
        plan.progress = progress

        if status:
            plan.status = status

        if progress == RepairProgress.COMPLETED and plan.status != RepairPlanStatus.COMPLETED:
            plan.status = RepairPlanStatus.COMPLETED
            plan.completed_at = datetime.now()

        db.commit()
        log_operation(db, current_user, OperationType.UPDATE, "repair_plan", plan_id, f"更新修复进度: {old_progress} -> {progress}")

    return RedirectResponse(url=f"/repair_plans/{plan_id}", status_code=status.HTTP_303_SEE_OTHER)


@app.post("/repair_plans/{plan_id}/images/upload")
async def upload_comparison_image(
    plan_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    image: UploadFile = File(...),
    image_type: str = Form(...),
    image_stage: str = Form(...),
    description: str = Form("")
):
    plan = db.query(RepairPlan).filter(RepairPlan.id == plan_id).first()
    if not plan:
        return RedirectResponse(url="/repair_plans", status_code=status.HTTP_303_SEE_OTHER)

    file_ext = os.path.splitext(image.filename)[1]
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(UPLOAD_DIR, unique_filename)

    with open(file_path, "wb") as buffer:
        content = await image.read()
        buffer.write(content)

    comparison_image = RepairComparisonImage(
        repair_plan_id=plan_id,
        image_path=f"/{file_path}",
        image_type=image_type,
        image_stage=image_stage,
        description=description,
        uploaded_by=current_user.id
    )
    db.add(comparison_image)
    db.commit()

    log_operation(db, current_user, OperationType.UPLOAD, "repair_comparison_image", comparison_image.id, f"上传修复对比图像: {plan.plan_name}")

    return RedirectResponse(url=f"/repair_plans/{plan_id}", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/repair_plans/{plan_id}/images/{image_id}/delete")
async def delete_comparison_image(
    plan_id: int,
    image_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    image = db.query(RepairComparisonImage).filter(
        RepairComparisonImage.id == image_id,
        RepairComparisonImage.repair_plan_id == plan_id
    ).first()
    if image:
        if os.path.exists(image.image_path.lstrip("/")):
            os.remove(image.image_path.lstrip("/"))
        db.delete(image)
        db.commit()
        log_operation(db, current_user, OperationType.DELETE, "repair_comparison_image", image_id, f"删除修复对比图像")

    return RedirectResponse(url=f"/repair_plans/{plan_id}", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/reviews", response_class=HTMLResponse)
async def list_reviews(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    status_filter: Optional[str] = None
):
    query = db.query(RepairPlan).filter(RepairPlan.status == RepairPlanStatus.IN_REVIEW)

    if status_filter:
        query = query.filter(RepairPlan.status == status_filter)

    plans = query.order_by(RepairPlan.submitted_at.desc()).all()

    return templates.TemplateResponse("reviews.html", {
        "request": request,
        "current_user": current_user,
        "plans": plans,
        "status_filter": status_filter
    })


@app.get("/reviews/{plan_id}", response_class=HTMLResponse)
async def review_page(
    plan_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    plan = db.query(RepairPlan).filter(RepairPlan.id == plan_id).first()
    if not plan:
        return RedirectResponse(url="/reviews", status_code=status.HTTP_303_SEE_OTHER)

    review_records = db.query(ReviewRecord).filter(ReviewRecord.repair_plan_id == plan_id).order_by(ReviewRecord.created_at.desc()).all()
    comparison_images = db.query(RepairComparisonImage).filter(RepairComparisonImage.repair_plan_id == plan_id).order_by(RepairComparisonImage.created_at.asc()).all()

    return templates.TemplateResponse("review_form.html", {
        "request": request,
        "current_user": current_user,
        "plan": plan,
        "review_records": review_records,
        "comparison_images": comparison_images,
        "today": date.today().isoformat()
    })


@app.post("/reviews/{plan_id}/submit")
async def submit_review(
    plan_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    review_opinion: str = Form(""),
    review_conclusion: str = Form(...),
    review_date: date = Form(...),
    is_returned: Optional[str] = Form("false"),
    return_reason: str = Form("")
):
    plan = db.query(RepairPlan).filter(RepairPlan.id == plan_id).first()
    if not plan:
        return RedirectResponse(url="/reviews", status_code=status.HTTP_303_SEE_OTHER)

    if review_date > date.today():
        review_records = db.query(ReviewRecord).filter(ReviewRecord.repair_plan_id == plan_id).order_by(ReviewRecord.created_at.desc()).all()
        comparison_images = db.query(RepairComparisonImage).filter(RepairComparisonImage.repair_plan_id == plan_id).order_by(RepairComparisonImage.created_at.asc()).all()
        return templates.TemplateResponse("review_form.html", {
            "request": request,
            "current_user": current_user,
            "plan": plan,
            "review_records": review_records,
            "comparison_images": comparison_images,
            "today": date.today().isoformat(),
            "error": "复核日期不能晚于当前日期"
        })

    returned = is_returned == "true"

    review = ReviewRecord(
        repair_plan_id=plan_id,
        reviewer_id=current_user.id,
        reviewer_name=current_user.full_name,
        review_opinion=review_opinion,
        review_conclusion=review_conclusion,
        review_date=review_date,
        is_returned=returned,
        return_reason=return_reason if returned else None
    )
    db.add(review)

    if returned:
        plan.status = RepairPlanStatus.REJECTED
        log_msg = f"复核退回: {return_reason}"
    elif review_conclusion == "通过":
        plan.status = RepairPlanStatus.APPROVED
        log_msg = "复核通过"
    else:
        plan.status = RepairPlanStatus.REJECTED
        log_msg = "复核不通过"

    db.commit()
    log_operation(db, current_user, OperationType.APPROVE if not returned else OperationType.REJECT, "repair_plan", plan_id, log_msg)

    return RedirectResponse(url="/reviews", status_code=status.HTTP_303_SEE_OTHER)


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
